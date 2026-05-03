"""Create comprehensive Excel report for UiPath Builder Agent code review."""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import json

def create_comprehensive_review_excel():
    """Generate comprehensive Excel workbook with all review data."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Colors
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Fonts
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)

    # === SHEET 1: Executive Summary ===
    ws1 = wb.create_sheet('Executive Summary')
    ws1['A1'] = 'UiPath Builder Agent - Code Review & Evaluation Report'
    ws1['A1'].font = title_font
    ws1['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'

    summary_data = [
        ['', ''],
        ['Metric', 'Value'],
        ['Overall Quality Score', '87/100'],
        ['Target Score', '95/100'],
        ['Gap to Target', '-8 points'],
        ['', ''],
        ['Test Results', '39/39 passing (100%)'],
        ['Code Coverage', '82%'],
        ['Python Files Reviewed', '23 files'],
        ['Total Lines of Code', '~1,800 lines'],
        ['', ''],
        ['Issues Found', ''],
        ['- Critical Issues', '4'],
        ['- Important Issues', '5'],
        ['- Suggestions', '4'],
        ['', ''],
        ['Status', 'PRODUCTION READY with fixes'],
        ['Estimated Fix Time', '2-3 days'],
    ]

    for i, row in enumerate(summary_data, start=4):
        ws1[f'A{i}'] = row[0]
        ws1[f'B{i}'] = row[1]
        if i == 5 or i == 11 or i == 16:
            ws1[f'A{i}'].font = bold_font
            ws1[f'B{i}'].font = bold_font

    # Style headers
    ws1['A5'].fill = header_fill
    ws1['B5'].fill = header_fill
    ws1['A5'].font = header_font
    ws1['B5'].font = header_font

    # Column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 40

    # === SHEET 2: Test Results ===
    ws2 = wb.create_sheet('Test Results')
    ws2['A1'] = 'Test Execution Results'
    ws2['A1'].font = title_font

    test_data = [
        ['', '', '', ''],
        ['Test Suite', 'Tests', 'Passed', 'Failed'],
        ['Integration Tests', '2', '2', '0'],
        ['Unit Tests - Bootstrap Flow', '29', '29', '0'],
        ['Unit Tests - Skill Discovery', '6', '6', '0'],
        ['Unit Tests - State Schema', '2', '2', '0'],
        ['', '', '', ''],
        ['TOTAL', '39', '39', '0'],
        ['', '', '', ''],
        ['Pass Rate', '100%', '', ''],
        ['Code Coverage', '82%', '', ''],
        ['Test Duration', '1.31 seconds', '', ''],
    ]

    for i, row in enumerate(test_data, start=3):
        for j, value in enumerate(row):
            cell = ws2.cell(row=i, column=j+1, value=value)
            if i == 4:  # Header row
                cell.fill = header_fill
                cell.font = header_font
            elif i == 10 or i == 12:  # Totals
                cell.font = bold_font

    # Apply green fill to passed column
    for i in range(5, 12):
        ws2.cell(row=i, column=3).fill = green_fill
        ws2.cell(row=i, column=4).fill = green_fill if ws2.cell(row=i, column=4).value == '0' else red_fill

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15

    # === SHEET 3: Code Quality Metrics ===
    ws3 = wb.create_sheet('Code Quality Metrics')
    ws3['A1'] = 'Code Quality Assessment'
    ws3['A1'].font = title_font

    quality_data = [
        ['', '', '', ''],
        ['Category', 'Score', 'Target', 'Status'],
        ['Plan Adherence', '95%', '95%', 'PASS'],
        ['Code Quality', '85%', '90%', 'NEEDS WORK'],
        ['Security', '70%', '95%', 'NEEDS WORK'],
        ['Error Handling', '75%', '90%', 'NEEDS WORK'],
        ['Type Safety', '80%', '90%', 'NEEDS WORK'],
        ['Test Coverage', '85%', '95%', 'NEEDS WORK'],
        ['Documentation', '85%', '90%', 'GOOD'],
        ['', '', '', ''],
        ['OVERALL', '87%', '95%', 'NEEDS FIXES'],
        ['', '', '', ''],
        ['Quality Checks', '', '', ''],
        ['Black Formatting', '20 files need formatting', '', 'ACTION REQUIRED'],
        ['Ruff Linting', '5 issues found', '', 'ACTION REQUIRED'],
        ['Mypy Type Checking', '5 type errors', '', 'ACTION REQUIRED'],
    ]

    for i, row in enumerate(quality_data, start=3):
        for j, value in enumerate(row):
            cell = ws3.cell(row=i, column=j+1, value=value)
            if i == 4:  # Header
                cell.fill = header_fill
                cell.font = header_font
            elif i == 12 or i == 14:  # Section headers
                cell.font = bold_font
            # Color code status
            if j == 3 and i > 4 and i < 11:
                if 'PASS' in str(value) or 'GOOD' in str(value):
                    cell.fill = green_fill
                elif 'NEEDS' in str(value):
                    cell.fill = yellow_fill
                elif 'FAIL' in str(value):
                    cell.fill = red_fill

    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 20

    # === SHEET 4: Critical Issues ===
    ws4 = wb.create_sheet('Critical Issues')
    ws4['A1'] = 'Critical Issues (Must Fix Before Production)'
    ws4['A1'].font = title_font

    issues_data = [
        ['', '', '', ''],
        ['#', 'Issue', 'File', 'Severity'],
        ['1', 'Security - Unvalidated User Input', 'cli/main.py:44,158', 'CRITICAL'],
        ['2', 'Architecture - Graph Entry Point Mismatch', 'agent/graph.py:84', 'CRITICAL'],
        ['3', 'Error Handling - Silent Skill Discovery Failures', 'agent/skill_discovery.py:85-107', 'IMPORTANT'],
        ['4', 'Security - Hardcoded AWS Credentials Risk', 'agent/nodes/*.py', 'IMPORTANT'],
        ['', '', '', ''],
        ['Details', '', '', ''],
        ['', '', '', ''],
        ['Issue #1: Unvalidated User Input', '', '', ''],
        ['Problem', 'User input passed directly to LLM without validation', '', ''],
        ['Impact', 'Prompt injection, resource exhaustion, DoS attacks', '', ''],
        ['Fix', 'Add input validation with max_length=5000, sanitize control chars', '', ''],
        ['', '', '', ''],
        ['Issue #2: Graph Entry Point', '', '', ''],
        ['Problem', 'Entry point is "ba" instead of "conversational" per spec', '', ''],
        ['Impact', 'Breaks dual-mode architecture, violates design spec', '', ''],
        ['Fix', 'Change builder.set_entry_point("conversational")', '', ''],
        ['', '', '', ''],
        ['Issue #3: Silent Failures', '', '', ''],
        ['Problem', 'Uses print() instead of logging, errors swallowed', '', ''],
        ['Impact', 'Hidden failures, difficult to debug in production', '', ''],
        ['Fix', 'Replace print() with logging.warning/error', '', ''],
        ['', '', '', ''],
        ['Issue #4: AWS Configuration', '', '', ''],
        ['Problem', 'Hardcoded region, no timeout/retry configuration', '', ''],
        ['Impact', 'Deployment inflexibility, hanging connections', '', ''],
        ['Fix', 'Use environment variables, add botocore Config with timeouts', '', ''],
    ]

    for i, row in enumerate(issues_data, start=3):
        for j, value in enumerate(row):
            cell = ws4.cell(row=i, column=j+1, value=value)
            if i == 4:  # Header
                cell.fill = header_fill
                cell.font = header_font
            elif 'CRITICAL' in str(value):
                cell.fill = red_fill
                cell.font = bold_font
            elif 'IMPORTANT' in str(value):
                cell.fill = yellow_fill
                cell.font = bold_font
            elif i in [10, 16, 22, 28]:  # Issue headers
                cell.font = bold_font
            elif j == 0 and i > 10:  # Labels
                cell.font = bold_font

    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 50
    ws4.column_dimensions['C'].width = 30
    ws4.column_dimensions['D'].width = 15

    # === SHEET 5: File Quality Breakdown ===
    ws5 = wb.create_sheet('File Quality Breakdown')
    ws5['A1'] = 'Quality Assessment by File'
    ws5['A1'].font = title_font

    file_data = [
        ['', '', '', '', ''],
        ['File', 'Lines', 'Coverage', 'Quality', 'Issues'],
        ['agent/state.py', '54', '100%', '95/100', '0'],
        ['agent/skill_discovery.py', '164', '90%', '85/100', '2'],
        ['agent/graph.py', '139', '84%', '75/100', '3'],
        ['agent/nodes/ba_persona.py', '106', '84%', '85/100', '1'],
        ['agent/nodes/sa_persona.py', '127', '69%', '85/100', '2'],
        ['agent/nodes/conversational.py', '88', '39%', '80/100', '2'],
        ['agent/nodes/developer_node.py', '169', '100%', '80/100', '3'],
        ['agent/nodes/hitl_node.py', '123', '80%', '90/100', '1'],
        ['agent/nodes/qa_node.py', '122', '84%', '90/100', '0'],
        ['agent/tools/skill_invoke.py', '123', '48%', '75/100', '4'],
        ['agent/prompts/constraints.py', '18', '100%', '95/100', '0'],
        ['cli/main.py', '184', '0%', '65/100', '6'],
        ['tests/unit/test_bootstrap_flow.py', '542', 'N/A', '95/100', '0'],
        ['tests/unit/test_skill_discovery.py', '142', 'N/A', '90/100', '0'],
        ['tests/unit/test_state.py', '45', 'N/A', '90/100', '0'],
        ['', '', '', '', ''],
        ['TOTALS', '1,800+', '82%', '87/100', '24'],
    ]

    for i, row in enumerate(file_data, start=3):
        for j, value in enumerate(row):
            cell = ws5.cell(row=i, column=j+1, value=value)
            if i == 4:  # Header
                cell.fill = header_fill
                cell.font = header_font
            elif i == 20:  # Totals
                cell.font = bold_font
            # Color code quality scores
            if j == 3 and i > 4 and i < 20:
                try:
                    score = int(str(value).split('/')[0])
                    if score >= 90:
                        cell.fill = green_fill
                    elif score >= 80:
                        cell.fill = yellow_fill
                    elif score < 80:
                        cell.fill = red_fill
                except:
                    pass

    ws5.column_dimensions['A'].width = 40
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 12
    ws5.column_dimensions['D'].width = 12
    ws5.column_dimensions['E'].width = 12

    # === SHEET 6: Recommendations ===
    ws6 = wb.create_sheet('Recommendations')
    ws6['A1'] = 'Actionable Recommendations'
    ws6['A1'].font = title_font

    rec_data = [
        ['', '', '', ''],
        ['Priority', 'Recommendation', 'Effort', 'Impact'],
        ['MUST FIX', 'Fix graph entry point to "conversational"', '30 min', 'HIGH'],
        ['MUST FIX', 'Add input validation in CLI', '1 hour', 'HIGH'],
        ['MUST FIX', 'Replace print() with proper logging', '1 hour', 'MEDIUM'],
        ['MUST FIX', 'Add LLM timeouts', '2 hours', 'HIGH'],
        ['MUST FIX', 'Add CLI tests', '4 hours', 'MEDIUM'],
        ['MUST FIX', 'Fix README (document Sprint 2)', '15 min', 'LOW'],
        ['', '', '', ''],
        ['SHOULD FIX', 'Add return type hints', '2 hours', 'MEDIUM'],
        ['SHOULD FIX', 'Remove code duplication (JSON parsing)', '1 hour', 'LOW'],
        ['SHOULD FIX', 'Add AWS SDK configuration', '2 hours', 'MEDIUM'],
        ['SHOULD FIX', 'Fix typos (Developement → Development)', '1 hour', 'LOW'],
        ['SHOULD FIX', 'Add UiPath version configuration', '1 hour', 'LOW'],
        ['SHOULD FIX', 'Add metrics/telemetry', '1 day', 'MEDIUM'],
        ['', '', '', ''],
        ['NICE TO HAVE', 'Add mypy to CI', '2 hours', 'LOW'],
        ['NICE TO HAVE', 'Add integration tests with real Bedrock', '1 day', 'MEDIUM'],
        ['NICE TO HAVE', 'Add performance tests', '2 days', 'LOW'],
        ['NICE TO HAVE', 'Add circuit breaker pattern', '1 day', 'MEDIUM'],
        ['', '', '', ''],
        ['Total Estimated Effort', '2-3 days for MUST FIX items', '', ''],
        ['Total Estimated Effort', '1 week for all items', '', ''],
    ]

    for i, row in enumerate(rec_data, start=3):
        for j, value in enumerate(row):
            cell = ws6.cell(row=i, column=j+1, value=value)
            if i == 4:  # Header
                cell.fill = header_fill
                cell.font = header_font
            elif 'MUST FIX' in str(value):
                cell.fill = red_fill
                cell.font = bold_font
            elif 'SHOULD FIX' in str(value):
                cell.fill = yellow_fill
                cell.font = bold_font
            elif 'NICE TO HAVE' in str(value):
                cell.fill = green_fill
            elif i >= 23:  # Totals
                cell.font = bold_font

    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 50
    ws6.column_dimensions['C'].width = 15
    ws6.column_dimensions['D'].width = 15

    # === Save workbook ===
    output_file = 'COMPREHENSIVE_CODE_REVIEW_REPORT.xlsx'
    wb.save(output_file)
    print(f'✅ Successfully created: {output_file}')
    print(f'   - Executive Summary')
    print(f'   - Test Results')
    print(f'   - Code Quality Metrics')
    print(f'   - Critical Issues')
    print(f'   - File Quality Breakdown')
    print(f'   - Recommendations')
    return output_file

if __name__ == '__main__':
    create_comprehensive_review_excel()
